"""
VFX Excel Output Generator
============================

Generates formatted Excel workbook with leads organized by persona tier
and company, with deal qualification status.

Output Structure:
- Sheet 1: Master Lead List (all leads, sorted by score)
- Sheet 2: By Persona Tier (grouped view)
- Sheet 3: Deal Qualification (company-level readiness)
- Sheet 4: Scoring Methodology
"""

import os
from datetime import datetime
from typing import List, Dict
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from data.roles import get_tier_label, get_tier_short, check_deal_qualification, classify_title


class VFXExcelExporter:
    """
    Exports VFX leads to a formatted Excel workbook.
    """
    
    COLUMN_WIDTHS = {
        'Name': 25,
        'Title': 35,
        'Company': 25,
        'Persona Tier': 18,
        'LinkedIn URL': 40,
        'Email': 30,
        'Location': 25,
        'Market': 12,
        'Score': 12,
        'Deal Status': 15,
        'Notes': 40,
    }
    
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    
    TIER_COLORS = {
        'Economic Buyer': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'Technical Champion': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        'Day-to-Day User': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'Procurement': PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        'Unclassified': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }
    
    SCORE_COLORS = {
        'high': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    
    def __init__(self, output_dir: str = 'output'):
        self.output_dir = output_dir
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    def export(self, leads: List[Dict]) -> str:
        """
        Export leads to Excel workbook.
        
        Args:
            leads: List of lead dicts with enriched data
            
        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"vfx_leads_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        self._create_master_sheet(wb, leads)
        self._create_tier_sheet(wb, leads)
        self._create_deal_qualification_sheet(wb, leads)
        self._create_methodology_sheet(wb)
        
        wb.save(filepath)
        
        # Also export CSV
        csv_path = filepath.replace('.xlsx', '.csv')
        self._export_csv(leads, csv_path)
        
        return filepath
    
    def _create_master_sheet(self, wb: Workbook, leads: List[Dict]):
        """Master lead list sorted by score."""
        ws = wb.create_sheet("Master Lead List")
        
        data = []
        for lead in leads:
            tier = lead.get('persona_tier') or classify_title(lead.get('title', ''))
            data.append({
                'Name': lead.get('name', ''),
                'Title': lead.get('title', ''),
                'Company': lead.get('company', ''),
                'Persona Tier': get_tier_label(tier),
                'LinkedIn URL': lead.get('linkedin_url', ''),
                'Email': lead.get('email', ''),
                'Location': lead.get('location', ''),
                'Market': lead.get('market', '').upper() if lead.get('market') else '',
                'Score': lead.get('score', 0),
            })
        
        df = pd.DataFrame(data)
        if not df.empty:
            df = df.sort_values('Score', ascending=False)
        
        self._write_dataframe_to_sheet(ws, df)
        self._apply_formatting(ws, df)
    
    def _create_tier_sheet(self, wb: Workbook, leads: List[Dict]):
        """Leads grouped by persona tier."""
        ws = wb.create_sheet("By Persona Tier")
        
        # Group leads by tier
        by_tier = defaultdict(list)
        for lead in leads:
            tier = lead.get('persona_tier') or classify_title(lead.get('title', ''))
            by_tier[tier].append(lead)
        
        current_row = 1
        tier_order = ['economic_buyer', 'technical_champion', 'day_to_day_user', 'procurement', 'unclassified']
        
        for tier_key in tier_order:
            tier_leads = by_tier.get(tier_key, [])
            if not tier_leads:
                continue
            
            label = get_tier_label(tier_key)
            
            # Tier header
            cell = ws.cell(row=current_row, column=1, value=f"{label} ({len(tier_leads)} leads)")
            cell.font = Font(bold=True, size=14)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1
            
            # Column headers
            columns = ['Name', 'Title', 'Company', 'LinkedIn URL', 'Email', 'Market', 'Score']
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            current_row += 1
            
            # Data rows
            sorted_leads = sorted(tier_leads, key=lambda x: x.get('score', 0), reverse=True)
            for lead in sorted_leads:
                ws.cell(row=current_row, column=1, value=lead.get('name', ''))
                ws.cell(row=current_row, column=2, value=lead.get('title', ''))
                ws.cell(row=current_row, column=3, value=lead.get('company', ''))
                ws.cell(row=current_row, column=4, value=lead.get('linkedin_url', ''))
                ws.cell(row=current_row, column=5, value=lead.get('email', ''))
                ws.cell(row=current_row, column=6, value=lead.get('market', '').upper() if lead.get('market') else '')
                ws.cell(row=current_row, column=7, value=lead.get('score', 0))
                current_row += 1
            
            current_row += 2  # Spacing between tiers
        
        # Column widths
        widths = [25, 35, 25, 40, 30, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    
    def _create_deal_qualification_sheet(self, wb: Workbook, leads: List[Dict]):
        """Company-level deal qualification view."""
        ws = wb.create_sheet("Deal Qualification")
        
        # Group leads by company and tier
        company_tiers = defaultdict(lambda: defaultdict(int))
        company_total = defaultdict(int)
        
        for lead in leads:
            company = lead.get('company', 'Unknown')
            tier = lead.get('persona_tier') or classify_title(lead.get('title', ''))
            company_tiers[company][tier] += 1
            company_total[company] += 1
        
        # Build rows
        data = []
        for company in sorted(company_tiers.keys()):
            tiers = company_tiers[company]
            qual = check_deal_qualification(tiers)
            
            data.append({
                'Company': company,
                'Total Leads': company_total[company],
                'Economic Buyers': tiers.get('economic_buyer', 0),
                'Technical Champions': tiers.get('technical_champion', 0),
                'Day-to-Day Users': tiers.get('day_to_day_user', 0),
                'Procurement': tiers.get('procurement', 0),
                'Coverage': qual['coverage'],
                'Deal Status': qual['summary'],
            })
        
        df = pd.DataFrame(data)
        if not df.empty:
            df = df.sort_values('Total Leads', ascending=False)
        
        self._write_dataframe_to_sheet(ws, df)
        
        # Format Deal Status column
        if not df.empty:
            status_col = list(df.columns).index('Deal Status') + 1
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=status_col)
                if cell.value == 'QUALIFIED':
                    cell.fill = self.SCORE_COLORS['high']
                    cell.font = Font(bold=True)
                elif cell.value and 'Missing' in str(cell.value):
                    cell.fill = self.SCORE_COLORS['low']
        
        # Column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = self.COLUMN_WIDTHS.get(col_name, 18)
        
        ws.freeze_panes = 'A2'
    
    def _create_methodology_sheet(self, wb: Workbook):
        """Scoring methodology explanation."""
        ws = wb.create_sheet("Scoring Methodology")
        
        ws['A1'] = "VFX Lead Scoring Methodology"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = "Persona Tiers (The Buyer Triangle)"
        ws['A3'].font = Font(bold=True, size=14)
        
        tiers = [
            ("Tier 1: Economic Buyer", "40%", "Head of Post, MD, EP, Head of VFX, COO", "Signs budget, approves vendors"),
            ("Tier 2: Technical Champion", "85 pts", "VFX Supe, CG Supe, Comp Supe, Pipeline TD", "Validates quality, workflow fit"),
            ("Tier 3: Day-to-Day User", "55 pts", "Senior Compositor, Lead Roto, Prep Supe", "Creates bottom-up demand"),
            ("Tier 4: Procurement", "40 pts", "Procurement Manager, Commercial Manager", "Later-stage only"),
        ]
        
        row = 4
        ws.cell(row=row, column=1, value="Tier").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Base Score").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Example Titles").font = Font(bold=True)
        ws.cell(row=row, column=4, value="Role").font = Font(bold=True)
        
        for tier, score, examples, role in tiers:
            row += 1
            ws.cell(row=row, column=1, value=tier)
            ws.cell(row=row, column=2, value=score)
            ws.cell(row=row, column=3, value=examples)
            ws.cell(row=row, column=4, value=role)
        
        row += 2
        ws['A' + str(row)] = "Deal Qualification Rule"
        ws['A' + str(row)].font = Font(bold=True, size=14)
        row += 1
        ws['A' + str(row)] = "Every serious opportunity should have:"
        row += 1
        ws['A' + str(row)] = "1. One Economic Buyer"
        row += 1
        ws['A' + str(row)] = "2. One Technical Champion"
        row += 1
        ws['A' + str(row)] = "3. One active Day-to-Day User group"
        row += 1
        ws['A' + str(row)] = "If any are missing, it's not a qualified deal."
        
        row += 2
        ws['A' + str(row)] = "Scoring Weights"
        ws['A' + str(row)].font = Font(bold=True, size=14)
        
        weights = [
            ("Persona Tier", "40%", "Which buyer triangle role they fill"),
            ("Seniority", "25%", "C-Suite > VP/Director > Supervisor > Lead > IC"),
            ("Company Relevance", "20%", "Blockbuster credits, company size"),
            ("Market Priority", "15%", "USA/UK (1.0x) > Canada (0.95x) > France (0.90x) > India (0.85x)"),
        ]
        
        row += 1
        ws.cell(row=row, column=1, value="Factor").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Weight").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Description").font = Font(bold=True)
        
        for factor, weight, desc in weights:
            row += 1
            ws.cell(row=row, column=1, value=factor)
            ws.cell(row=row, column=2, value=weight)
            ws.cell(row=row, column=3, value=desc)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """Write a dataframe to a worksheet."""
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _apply_formatting(self, ws, df: pd.DataFrame):
        """Apply formatting to the master sheet."""
        # Column widths
        for col_idx, column in enumerate(df.columns, 1):
            width = self.COLUMN_WIDTHS.get(column, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Color-code persona tier column
        if 'Persona Tier' in df.columns:
            tier_col = list(df.columns).index('Persona Tier') + 1
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=tier_col)
                fill = self.TIER_COLORS.get(cell.value)
                if fill:
                    cell.fill = fill
        
        # Color-code score column
        if 'Score' in df.columns:
            score_col = list(df.columns).index('Score') + 1
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=score_col)
                try:
                    score = int(cell.value or 0)
                    if score >= 75:
                        cell.fill = self.SCORE_COLORS['high']
                    elif score >= 50:
                        cell.fill = self.SCORE_COLORS['medium']
                    else:
                        cell.fill = self.SCORE_COLORS['low']
                except (ValueError, TypeError):
                    pass
        
        ws.freeze_panes = 'A2'
        if len(df) > 0:
            ws.auto_filter.ref = ws.dimensions
    
    def _export_csv(self, leads: List[Dict], csv_path: str):
        """Export leads to CSV."""
        data = []
        for lead in leads:
            tier = lead.get('persona_tier') or classify_title(lead.get('title', ''))
            data.append({
                'Name': lead.get('name', ''),
                'Title': lead.get('title', ''),
                'Company': lead.get('company', ''),
                'Persona Tier': get_tier_label(tier),
                'LinkedIn URL': lead.get('linkedin_url', ''),
                'Email': lead.get('email', ''),
                'Location': lead.get('location', ''),
                'Market': lead.get('market', ''),
                'Score': lead.get('score', 0),
            })
        
        df = pd.DataFrame(data)
        if not df.empty:
            df = df.sort_values('Score', ascending=False)
        df.to_csv(csv_path, index=False)
