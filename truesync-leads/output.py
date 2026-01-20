"""
Excel Output Generator
======================

Generates formatted Excel workbook with leads organized by market.

Output Structure:
- Sheet 1: Master Lead List (all leads, sortable)
- Sheet 2: Spain Breakdown
- Sheet 3: Korea Breakdown
- Sheet 4: France Breakdown
- Sheet 5: Catalog Intelligence
- Sheet 6: Scoring Methodology
"""

import os
from datetime import datetime
from typing import List, Optional
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from enrichers.context import CatalogContextEnricher


class ExcelExporter:
    """
    Exports leads to a formatted Excel workbook.
    """
    
    # Column widths for consistent formatting
    COLUMN_WIDTHS = {
        'Name': 25,
        'Title': 35,
        'Company': 25,
        'Company Type': 15,
        'LinkedIn URL': 40,
        'Email': 30,
        'Market': 12,
        'Priority Score': 15,
        'Catalog Context': 50,
        'Created': 20,
        'Updated': 20
    }
    
    # Header style
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    
    # Score-based row colors
    SCORE_COLORS = {
        'high': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # Green
        'medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # Yellow
        'low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # Red
    }
    
    def __init__(self, output_dir: str = 'output'):
        self.output_dir = output_dir
        self.catalog_enricher = CatalogContextEnricher()
        
        # Ensure output directory exists
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    def export(self, leads: List, companies: List) -> str:
        """
        Export leads to Excel workbook.
        
        Args:
            leads: List of Lead objects
            companies: List of Company objects
            
        Returns:
            Path to the generated Excel file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"truesync_leads_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create sheets
        self._create_master_sheet(wb, leads)
        # Tier 1 - Primary Markets
        self._create_market_sheet(wb, leads, 'UK', 'uk')
        self._create_market_sheet(wb, leads, 'USA', 'usa')
        self._create_market_sheet(wb, leads, 'Spain', 'spain')
        # Tier 2 - Expansion Markets
        self._create_market_sheet(wb, leads, 'Germany', 'germany')
        self._create_market_sheet(wb, leads, 'France', 'france')
        self._create_market_sheet(wb, leads, 'Korea', 'korea')
        self._create_catalog_sheet(wb, companies)
        self._create_methodology_sheet(wb)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
    
    def _create_master_sheet(self, wb: Workbook, leads: List):
        """Create the master lead list sheet."""
        ws = wb.create_sheet("Master Lead List")
        
        # Convert leads to dataframe
        data = []
        for lead in leads:
            catalog_context = ''
            if lead.company:
                catalog_context = self.catalog_enricher.get_context_for_lead(
                    lead.company.name, lead.market
                )
            
            data.append({
                'Name': lead.name,
                'Title': lead.title,
                'Company': lead.company.name if lead.company else '',
                'Company Type': lead.company.type if lead.company else '',
                'LinkedIn URL': lead.linkedin_url or '',
                'Email': lead.email or '',
                'Market': lead.market.title() if lead.market else '',
                'Priority Score': lead.priority_score or 0,
                'Catalog Context': catalog_context,
                'Created': lead.created_at.strftime('%Y-%m-%d') if lead.created_at else '',
                'Updated': lead.updated_at.strftime('%Y-%m-%d') if lead.updated_at else ''
            })
        
        df = pd.DataFrame(data)
        
        # Sort by priority score descending
        if not df.empty:
            df = df.sort_values('Priority Score', ascending=False)
        
        self._write_dataframe_to_sheet(ws, df)
        self._apply_formatting(ws, df)
    
    def _create_market_sheet(
        self,
        wb: Workbook,
        leads: List,
        sheet_name: str,
        market: str
    ):
        """Create a market-specific sheet."""
        ws = wb.create_sheet(sheet_name)
        
        # Filter leads by market
        market_leads = [l for l in leads if l.market == market]
        
        # Group by company
        companies = {}
        for lead in market_leads:
            company_name = lead.company.name if lead.company else 'Unknown'
            if company_name not in companies:
                companies[company_name] = []
            companies[company_name].append(lead)
        
        # Write each company section
        current_row = 1
        for company_name, company_leads in companies.items():
            # Company header
            ws.cell(row=current_row, column=1, value=company_name)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=6
            )
            current_row += 1
            
            # Lead data
            data = []
            for lead in sorted(company_leads, key=lambda x: x.priority_score or 0, reverse=True):
                data.append({
                    'Name': lead.name,
                    'Title': lead.title,
                    'LinkedIn URL': lead.linkedin_url or '',
                    'Email': lead.email or '',
                    'Priority Score': lead.priority_score or 0
                })
            
            df = pd.DataFrame(data)
            
            # Write headers
            for col_idx, column in enumerate(df.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=column)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            
            current_row += 1
            
            # Write data
            for _, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            # Add spacing between companies
            current_row += 2
        
        # Adjust column widths
        for col_idx in range(1, 6):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    def _create_catalog_sheet(self, wb: Workbook, companies: List):
        """Create the catalog intelligence sheet."""
        ws = wb.create_sheet("Catalog Intelligence")
        
        data = []
        for company in companies:
            catalog_data = self.catalog_enricher.get_full_catalog_data(
                company.name, company.market
            )
            
            # Get top titles
            movies = catalog_data.get('tmdb_movies', [])
            tv_shows = catalog_data.get('tmdb_tv_shows', [])
            
            top_movies = ', '.join([m['title'] for m in movies[:3]]) if movies else 'N/A'
            top_shows = ', '.join([s['title'] for s in tv_shows[:3]]) if tv_shows else 'N/A'
            
            data.append({
                'Company': company.name,
                'Market': company.market.title(),
                'Type': company.type,
                'Catalog Size': company.catalog_size or 'Unknown',
                'Non-English Count': catalog_data.get('non_english_count', 0),
                'Top Movies': top_movies,
                'Top TV Shows': top_shows,
                'Notes': company.catalog_notes or ''
            })
        
        df = pd.DataFrame(data)
        self._write_dataframe_to_sheet(ws, df)
        
        # Apply formatting
        for col_idx, column in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30
    
    def _create_methodology_sheet(self, wb: Workbook):
        """Create the scoring methodology sheet."""
        ws = wb.create_sheet("Scoring Methodology")
        
        # Title
        ws['A1'] = "TrueSync Lead Scoring Methodology"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Overview
        ws['A3'] = "Overview"
        ws['A3'].font = Font(bold=True, size=14)
        ws['A4'] = "Leads are scored 1-100 based on their potential value for TrueSync sales."
        ws['A5'] = "Higher scores indicate higher priority for outreach."
        
        # Scoring factors
        ws['A7'] = "Scoring Factors"
        ws['A7'].font = Font(bold=True, size=14)
        
        factors = [
            ("Role Relevance", "25%", "How relevant is their role to content distribution/licensing decisions"),
            ("Catalog Volume", "25%", "Size of their company's non-English catalog"),
            ("Market Priority", "20%", "Spain (1.0x) > Korea (0.95x) > France (0.90x)"),
            ("Decision Authority", "15%", "Seniority level (Chief/VP/Director/Manager)"),
            ("Company Type", "15%", "Distributor (1.0x) > Producer (0.9x) > Platform (0.85x)")
        ]
        
        row = 8
        ws.cell(row=row, column=1, value="Factor").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Weight").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Description").font = Font(bold=True)
        
        for factor, weight, desc in factors:
            row += 1
            ws.cell(row=row, column=1, value=factor)
            ws.cell(row=row, column=2, value=weight)
            ws.cell(row=row, column=3, value=desc)
        
        # High priority roles
        ws.cell(row=row+2, column=1, value="High Priority Roles").font = Font(bold=True, size=14)
        
        roles = [
            "Head of International Sales/Distribution",
            "VP/SVP International Licensing",
            "Director of Content Sales",
            "Head of Global Distribution",
            "Chief Content Officer"
        ]
        
        for i, role in enumerate(roles, 1):
            ws.cell(row=row+2+i, column=1, value=f"• {role}")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """Write a dataframe to a worksheet."""
        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _apply_formatting(self, ws, df: pd.DataFrame):
        """Apply formatting to a worksheet."""
        # Adjust column widths
        for col_idx, column in enumerate(df.columns, 1):
            width = self.COLUMN_WIDTHS.get(column, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Apply score-based coloring
        if 'Priority Score' in df.columns:
            score_col = list(df.columns).index('Priority Score') + 1
            
            for row_idx in range(2, len(df) + 2):
                score_cell = ws.cell(row=row_idx, column=score_col)
                try:
                    score = int(score_cell.value or 0)
                    if score >= 75:
                        fill = self.SCORE_COLORS['high']
                    elif score >= 50:
                        fill = self.SCORE_COLORS['medium']
                    else:
                        fill = self.SCORE_COLORS['low']
                    score_cell.fill = fill
                except (ValueError, TypeError):
                    pass
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add auto-filter
        if len(df) > 0:
            ws.auto_filter.ref = ws.dimensions


# Example usage
if __name__ == '__main__':
    exporter = ExcelExporter()
    
    # This would normally use real leads from the database
    print("Excel exporter ready. Use export() method with leads and companies.")
