#!/usr/bin/env python3
"""
TrueSync Accounts Pipeline
==========================

Enriches all target companies from data/companies.py with:
- TMDb catalog data (title counts, top shows, popularity)
- Google Search fallback for missing data
- Contact counts from leads database

Output Schema:
- Accounts (company name)
- Region (market)
- No. of titles (from TMDb or Google Search)
- Type of company (Producer/Distributor/Platform)
- No. of contacts at company (from leads DB)

Usage:
    python execution/accounts_pipeline.py
    python execution/accounts_pipeline.py --skip-db
    python execution/accounts_pipeline.py --skip-cache

Pipeline Flow:
1. Load all companies from data/companies.py
2. Enrich with TMDb data (title count, top shows)
3. Fallback to Google Search if TMDb data insufficient
4. Count contacts from leads database
5. Export to database (accounts table) and Excel
"""

import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import centralized utilities and data
from data.companies import COMPANIES, get_company_domains
from scrapers.catalog import TMDbClient
from utils import setup_logging, get_logger, retry_with_backoff, APICache


# Initialize cache for API results
_api_cache = None


def get_api_cache() -> APICache:
    """Get or create the API cache instance."""
    global _api_cache
    if _api_cache is None:
        _api_cache = APICache(cache_dir='.cache', ttl_hours=24)
    return _api_cache


# ============================================================================
# STEP 1: Load Companies from Knowledge Base
# ============================================================================

def load_companies_from_knowledge_base() -> list:
    """
    Load all target companies from data/companies.py.
    
    Returns:
        List of company dictionaries with name, type, market, catalog info
    """
    logger = get_logger()
    
    logger.info("=" * 60)
    logger.info("📂 STEP 1: Load Companies from Knowledge Base")
    logger.info("=" * 60)
    
    companies = []
    for company in COMPANIES:
        companies.append({
            'name': company.get('name', ''),
            'type': company.get('type', ''),
            'region': company.get('market', ''),
            'linkedin_url': company.get('linkedin_url', ''),
            'catalog_size_notes': company.get('catalog_size', ''),
            'catalog_notes': company.get('catalog_notes', ''),
        })
    
    logger.info(f"   Loaded {len(companies)} companies from data/companies.py")
    
    # Log by market
    markets = {}
    for c in companies:
        m = c.get('region', 'unknown')
        markets[m] = markets.get(m, 0) + 1
    
    for market, count in sorted(markets.items()):
        logger.info(f"   {market.upper()}: {count} companies")
    
    return companies


# ============================================================================
# STEP 2: Enrich with TMDb Data
# ============================================================================

def enrich_with_tmdb(companies: list, use_cache: bool = True) -> list:
    """
    Enrich companies with TMDb catalog data.
    
    Args:
        companies: List of company dictionaries
        use_cache: Whether to use cached results
        
    Returns:
        Companies with added TMDb data (num_titles, top_shows, tmdb_id)
    """
    logger = get_logger()
    cache = get_api_cache() if use_cache else None
    
    logger.info("=" * 60)
    logger.info("🎬 STEP 2: TMDb Enrichment")
    logger.info("=" * 60)
    
    tmdb = TMDbClient()
    
    if not tmdb.api_key and not tmdb.access_token:
        logger.warning("   ⚠ TMDb credentials not configured. Set TMDB_ACCESS_TOKEN or TMDB_API_KEY")
        logger.info("   Will rely on Google Search fallback for all companies")
        return companies
    
    enriched_count = 0
    
    for company in companies:
        name = company.get('name', '')
        market = company.get('region', 'usa')
        
        # Check cache first
        cache_key = f"tmdb_company_{name.lower().replace(' ', '_')}"
        cached_data = None
        
        if cache:
            cached_data = cache.get(cache_key)
            if cached_data:
                logger.debug(f"   📦 Cache hit: {name}")
                company.update(cached_data)
                enriched_count += 1
                continue
        
        # Query TMDb
        try:
            show_data = tmdb.get_top_shows_formatted(name, market)
            
            if show_data.get('total_catalog', 0) > 0:
                company['num_titles'] = show_data.get('total_catalog', 0)
                company['top_shows'] = show_data.get('top_shows', '')
                company['popularity_score'] = max(
                    [s.get('popularity', 0) for s in show_data.get('show_details', [])] or [0]
                )
                company['tmdb_id'] = tmdb.search_company(name)
                company['data_source'] = 'tmdb'
                enriched_count += 1
                
                # Cache successful result
                if cache:
                    cache.set(cache_key, {
                        'num_titles': company['num_titles'],
                        'top_shows': company['top_shows'],
                        'popularity_score': company['popularity_score'],
                        'tmdb_id': company['tmdb_id'],
                        'data_source': 'tmdb'
                    })
                
                logger.info(f"   ✓ {name}: {company['num_titles']} titles")
            else:
                logger.info(f"   ○ {name}: No TMDb data (will try Google Search)")
                company['data_source'] = 'pending_google'
                
        except Exception as e:
            logger.warning(f"   ⚠ {name}: TMDb error - {str(e)[:50]}")
            company['data_source'] = 'pending_google'
    
    logger.info(f"   ✅ TMDb enriched: {enriched_count}/{len(companies)} companies")
    
    return companies


# ============================================================================
# STEP 3: Google Search Fallback
# ============================================================================

@retry_with_backoff(max_retries=3, base_delay=5.0, exceptions=(Exception,))
def _run_google_search(client, query: str) -> list:
    """Run Google search with retry logic."""
    actor_id = "apify/google-search-scraper"
    run = client.actor(actor_id).call(
        run_input={
            "queries": query,
            "maxPagesPerQuery": 1,
            "resultsPerPage": 5,
        },
        timeout_secs=120,
        memory_mbytes=512
    )
    return list(client.dataset(run["defaultDatasetId"]).iterate_items())


def enrich_with_google_search(companies: list, use_cache: bool = True) -> list:
    """
    Fallback: Use Google Search to find title counts for companies
    where TMDb data was insufficient.
    
    Args:
        companies: List of company dictionaries
        use_cache: Whether to use cached results
        
    Returns:
        Companies with Google Search enrichment for missing data
    """
    logger = get_logger()
    cache = get_api_cache() if use_cache else None
    
    # Find companies needing Google Search fallback
    needs_search = [c for c in companies if c.get('data_source') == 'pending_google']
    
    if not needs_search:
        logger.info("   ✅ All companies have TMDb data, skipping Google Search")
        return companies
    
    logger.info("=" * 60)
    logger.info("🔍 STEP 3: Google Search Fallback")
    logger.info("=" * 60)
    logger.info(f"   Companies needing Google Search: {len(needs_search)}")
    
    token = os.getenv('APIFY_TOKEN')
    if not token:
        logger.warning("   ⚠ APIFY_TOKEN not found. Cannot run Google Search fallback.")
        # Mark remaining as manual and use catalog_size_notes
        for company in needs_search:
            company['data_source'] = 'catalog_notes'
            # Try to extract number from catalog_size_notes
            notes = company.get('catalog_size_notes', '')
            import re
            numbers = re.findall(r'(\d+,?\d*)', notes.replace(',', ''))
            if numbers:
                company['num_titles'] = int(numbers[0].replace(',', ''))
        return companies
    
    from apify_client import ApifyClient
    client = ApifyClient(token)
    
    enriched_count = 0
    
    for company in needs_search:
        name = company.get('name', '')
        
        # Check cache
        cache_key = f"google_catalog_{name.lower().replace(' ', '_')}"
        if cache:
            cached = cache.get(cache_key)
            if cached:
                company.update(cached)
                enriched_count += 1
                logger.debug(f"   📦 Cache hit: {name}")
                continue
        
        # Search Google
        query = f'"{name}" production company catalog titles films TV shows'
        
        try:
            results = _run_google_search(client, query)
            
            # Parse results for title information
            num_titles = 0
            top_shows = []
            
            for result in results:
                for organic in result.get('organicResults', []):
                    snippet = organic.get('description', '') + ' ' + organic.get('title', '')
                    
                    # Look for numbers near "titles", "films", "shows", "hours"
                    import re
                    patterns = [
                        r'(\d+,?\d*)\+?\s*(?:titles|films|movies|shows|hours|productions)',
                        r'(?:catalog|library|archive)\s*(?:of\s*)?(\d+,?\d*)',
                    ]
                    
                    for pattern in patterns:
                        matches = re.findall(pattern, snippet.lower())
                        if matches:
                            for match in matches:
                                num = int(match.replace(',', ''))
                                if num > num_titles:
                                    num_titles = num
            
            if num_titles > 0:
                company['num_titles'] = num_titles
                company['data_source'] = 'google_search'
                enriched_count += 1
                
                # Cache result
                if cache:
                    cache.set(cache_key, {
                        'num_titles': num_titles,
                        'data_source': 'google_search'
                    })
                
                logger.info(f"   ✓ {name}: {num_titles} titles (Google)")
            else:
                # Fall back to catalog notes
                company['data_source'] = 'catalog_notes'
                notes = company.get('catalog_size_notes', '')
                numbers = re.findall(r'(\d+,?\d*)', notes.replace(',', ''))
                if numbers:
                    company['num_titles'] = int(numbers[0].replace(',', ''))
                    logger.info(f"   ○ {name}: {company['num_titles']} titles (from notes)")
                else:
                    logger.info(f"   ○ {name}: No title count found")
                    
        except Exception as e:
            logger.warning(f"   ⚠ {name}: Google Search error - {str(e)[:50]}")
            company['data_source'] = 'error'
    
    logger.info(f"   ✅ Google Search enriched: {enriched_count}/{len(needs_search)} companies")
    
    return companies


# ============================================================================
# STEP 4: Count Contacts from Leads Database
# ============================================================================

def count_contacts_from_database(companies: list) -> list:
    """
    Count existing leads per company from the database.
    
    Args:
        companies: List of company dictionaries
        
    Returns:
        Companies with num_contacts populated
    """
    logger = get_logger()
    
    logger.info("=" * 60)
    logger.info("👥 STEP 4: Count Contacts from Leads Database")
    logger.info("=" * 60)
    
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        logger.warning("   ⚠ DATABASE_URL not found. Cannot count contacts from database.")
        # Set all to 0
        for company in companies:
            company['num_contacts'] = 0
        return companies
    
    try:
        from db.connection import get_engine
        from sqlalchemy import text
        
        engine = get_engine()
        
        # Query contact counts grouped by company name
        # Using fuzzy matching on company name
        with engine.connect() as conn:
            # Get all leads with company info
            result = conn.execute(text("""
                SELECT 
                    COALESCE(c.name, 'Unknown') as company_name,
                    COUNT(l.id) as contact_count
                FROM leads l
                LEFT JOIN companies c ON l.company_id = c.id
                GROUP BY c.name
            """))
            
            db_counts = {row[0]: row[1] for row in result}
        
        logger.info(f"   Found contact counts for {len(db_counts)} companies in database")
        
        # Match to our companies (fuzzy matching)
        total_contacts = 0
        for company in companies:
            name = company.get('name', '')
            
            # Try exact match first
            if name in db_counts:
                company['num_contacts'] = db_counts[name]
                total_contacts += company['num_contacts']
            else:
                # Try partial match
                matched = False
                for db_name, count in db_counts.items():
                    if db_name and (name.lower() in db_name.lower() or db_name.lower() in name.lower()):
                        company['num_contacts'] = count
                        total_contacts += count
                        matched = True
                        break
                
                if not matched:
                    company['num_contacts'] = 0
        
        logger.info(f"   ✅ Total contacts across all companies: {total_contacts}")
        
    except Exception as e:
        logger.warning(f"   ⚠ Database error: {e}")
        for company in companies:
            company['num_contacts'] = 0
    
    return companies


# ============================================================================
# STEP 5: Export to Database and Excel
# ============================================================================

def export_to_database(companies: list) -> int:
    """
    Push enriched company data to the accounts table.
    
    Args:
        companies: List of enriched company dictionaries
        
    Returns:
        Number of records inserted/updated
    """
    logger = get_logger()
    
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        logger.warning("   ⚠ DATABASE_URL not found. Skipping database export.")
        return 0
    
    try:
        from db.connection import get_engine
        from db.models import Account
        from sqlalchemy.orm import Session
        from sqlalchemy import select
        
        engine = get_engine()
        
        inserted = 0
        updated = 0
        
        with Session(engine) as session:
            for company in companies:
                name = company.get('name', '')
                
                # Check for existing account
                stmt = select(Account).where(Account.name == name)
                existing = session.execute(stmt).scalar_one_or_none()
                
                if existing:
                    # Update existing
                    existing.region = company.get('region', existing.region)
                    existing.num_titles = company.get('num_titles', existing.num_titles)
                    existing.company_type = company.get('type', existing.company_type)
                    existing.num_contacts = company.get('num_contacts', existing.num_contacts)
                    existing.top_shows = company.get('top_shows', existing.top_shows)
                    existing.tmdb_id = company.get('tmdb_id', existing.tmdb_id)
                    existing.popularity_score = company.get('popularity_score', existing.popularity_score)
                    existing.data_source = company.get('data_source', existing.data_source)
                    updated += 1
                else:
                    # Insert new
                    import uuid
                    new_account = Account(
                        id=str(uuid.uuid4()),
                        name=name,
                        region=company.get('region', ''),
                        num_titles=company.get('num_titles', 0),
                        company_type=company.get('type', ''),
                        num_contacts=company.get('num_contacts', 0),
                        top_shows=company.get('top_shows', ''),
                        tmdb_id=company.get('tmdb_id'),
                        popularity_score=company.get('popularity_score', 0),
                        data_source=company.get('data_source', ''),
                    )
                    session.add(new_account)
                    inserted += 1
            
            session.commit()
        
        logger.info(f"   ✅ Database: {inserted} inserted, {updated} updated")
        return inserted + updated
        
    except Exception as e:
        logger.error(f"   ❌ Database error: {e}")
        return 0


def export_to_excel(companies: list) -> str:
    """
    Export enriched company data to styled Excel workbook.
    
    Args:
        companies: List of enriched company dictionaries
        
    Returns:
        Path to exported Excel file
    """
    logger = get_logger()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f"output/accounts_{timestamp}.xlsx"
    
    Path("output").mkdir(exist_ok=True)
    
    # Prepare data for main sheet
    df_data = []
    for company in companies:
        df_data.append({
            'Accounts': company.get('name', ''),
            'Region': company.get('region', '').upper(),
            'No. of titles': company.get('num_titles', 0),
            'Type of company': company.get('type', ''),
            'No. of contacts': company.get('num_contacts', 0),
        })
    
    df = pd.DataFrame(df_data)
    df = df.sort_values(['Region', 'No. of titles'], ascending=[True, False])
    
    # Create styled workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    column_widths = {
        'A': 35,  # Accounts
        'B': 12,  # Region
        'C': 15,  # No. of titles
        'D': 20,  # Type of company
        'E': 18,  # No. of contacts
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions
    
    # Create "Show Details" sheet
    ws_details = wb.create_sheet("Show Details")
    
    # Prepare details data
    details_data = []
    for company in companies:
        if company.get('top_shows'):
            details_data.append({
                'Account': company.get('name', ''),
                'Top Shows': company.get('top_shows', ''),
                'Popularity Score': company.get('popularity_score', 0),
                'Data Source': company.get('data_source', ''),
            })
    
    if details_data:
        df_details = pd.DataFrame(details_data)
        
        # Write headers
        for col_idx, header in enumerate(df_details.columns, 1):
            cell = ws_details.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, row in enumerate(df_details.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border
        
        # Set column widths
        ws_details.column_dimensions['A'].width = 30
        ws_details.column_dimensions['B'].width = 80
        ws_details.column_dimensions['C'].width = 18
        ws_details.column_dimensions['D'].width = 15
        
        ws_details.freeze_panes = 'A2'
    
    # Save
    wb.save(output_path)
    
    # Also save CSV
    csv_path = output_path.replace('.xlsx', '.csv')
    df.to_csv(csv_path, index=False)
    
    logger.info(f"   📊 Excel: {output_path}")
    logger.info(f"   📊 CSV: {csv_path}")
    
    return output_path


# ============================================================================
# EXPORT FROM DATABASE (Sync CSV with DB)
# ============================================================================

def export_accounts_from_database() -> str:
    """
    Export accounts directly from database to CSV/Excel.
    
    This ensures output files are always in sync with database state.
    
    Returns:
        Path to exported Excel file, or None if database not available
    """
    logger = get_logger()
    
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        logger.warning("   ⚠ DATABASE_URL not found. Cannot export from database.")
        return None
    
    try:
        from db.connection import get_engine
        from db.models import Account
        from sqlalchemy.orm import Session
        from sqlalchemy import select
        
        engine = get_engine()
        
        with Session(engine) as session:
            # Fetch all accounts from database
            stmt = select(Account).order_by(Account.region, Account.num_titles.desc())
            accounts = session.execute(stmt).scalars().all()
            
            if not accounts:
                logger.warning("   ⚠ No accounts found in database")
                return None
            
            # Convert to company dicts for export_to_excel
            companies = []
            for acc in accounts:
                companies.append({
                    'name': acc.name,
                    'region': acc.region,
                    'num_titles': acc.num_titles or 0,
                    'type': acc.company_type,
                    'num_contacts': acc.num_contacts or 0,
                    'top_shows': acc.top_shows,
                    'popularity_score': acc.popularity_score or 0,
                    'data_source': acc.data_source,
                })
            
            logger.info(f"   📦 Loaded {len(companies)} accounts from database")
            
            # Export to Excel/CSV
            output_path = export_to_excel(companies)
            
            logger.info(f"   ✅ Exported {len(companies)} accounts from database to CSV/Excel")
            
            return output_path
            
    except Exception as e:
        logger.error(f"   ❌ Database export error: {e}")
        return None


def export_leads_from_database() -> str:
    """
    Export leads directly from database to CSV.
    
    Returns:
        Path to exported CSV file, or None if database not available
    """
    logger = get_logger()
    
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        logger.warning("   ⚠ DATABASE_URL not found. Cannot export leads from database.")
        return None
    
    try:
        from db.connection import get_engine
        from db.models import Lead, Company
        from sqlalchemy.orm import Session
        from sqlalchemy import select
        
        engine = get_engine()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        Path("output").mkdir(exist_ok=True)
        
        with Session(engine) as session:
            # Fetch all leads with company info
            stmt = select(Lead).order_by(Lead.priority_score.desc())
            leads = session.execute(stmt).scalars().all()
            
            if not leads:
                logger.warning("   ⚠ No leads found in database")
                return None
            
            # Convert to list of dicts
            leads_data = []
            for lead in leads:
                leads_data.append({
                    'Name': lead.name,
                    'Title': lead.title,
                    'Company': lead.company.name if lead.company else '',
                    'Company Type': lead.company.type if lead.company else '',
                    'LinkedIn URL': lead.linkedin_url,
                    'Email': lead.email,
                    'Market': lead.market.upper() if lead.market else '',
                    'Priority Score': lead.priority_score,
                    'Catalog Context': lead.catalog_context,
                })
            
            # Export to CSV
            df = pd.DataFrame(leads_data)
            csv_path = f"output/leads_from_db_{timestamp}.csv"
            df.to_csv(csv_path, index=False)
            
            logger.info(f"   📊 Exported {len(leads_data)} leads to {csv_path}")
            
            # Also export by market
            market_dir = Path("output/by_market")
            market_dir.mkdir(exist_ok=True)
            
            for market in df['Market'].unique():
                if market:
                    market_df = df[df['Market'] == market]
                    market_path = market_dir / f"leads_{market.lower()}.csv"
                    market_df.to_csv(market_path, index=False)
                    logger.info(f"   📊 {market}: {len(market_df)} leads → {market_path}")
            
            return csv_path
            
    except Exception as e:
        logger.error(f"   ❌ Leads export error: {e}")
        return None


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="TrueSync Accounts Pipeline - Enrich companies with title counts and contact data")
    parser.add_argument("--skip-db", action="store_true", help="Skip database push")
    parser.add_argument("--skip-cache", action="store_true", help="Skip API result caching")
    parser.add_argument("--clear-cache", action="store_true", help="Clear API cache before running")
    parser.add_argument("--export-from-db", action="store_true", help="Export accounts and leads from database to CSV (skip enrichment)")
    parser.add_argument("--export-leads", action="store_true", help="Also export leads from database")
    
    args = parser.parse_args()
    
    # Set up logging
    logger = setup_logging(output_dir='output', log_to_file=True)
    use_cache = not args.skip_cache
    
    # Handle --export-from-db mode (skip enrichment, just export from DB)
    if args.export_from_db:
        logger.info("=" * 70)
        logger.info("📤 EXPORT FROM DATABASE MODE")
        logger.info("=" * 70)
        
        accounts_path = export_accounts_from_database()
        
        if args.export_leads:
            leads_path = export_leads_from_database()
        
        logger.info("=" * 70)
        logger.info("✅ DATABASE EXPORT COMPLETE")
        logger.info("=" * 70)
        return
    
    logger.info("=" * 70)
    logger.info("🏢 TRUESYNC ACCOUNTS PIPELINE")
    logger.info("   - Enriches all target companies with catalog data")
    logger.info("   - TMDb primary, Google Search fallback")
    logger.info("   - Counts contacts from leads database")
    logger.info("   - Outputs to database and styled Excel")
    logger.info("=" * 70)
    
    # Clear cache if requested
    if args.clear_cache:
        cache = get_api_cache()
        cache.clear()
        logger.info("   Cache cleared")
    
    # STEP 1: Load companies
    companies = load_companies_from_knowledge_base()
    
    # STEP 2: TMDb enrichment
    companies = enrich_with_tmdb(companies, use_cache=use_cache)
    
    # STEP 3: Google Search fallback
    companies = enrich_with_google_search(companies, use_cache=use_cache)
    
    # STEP 4: Count contacts
    companies = count_contacts_from_database(companies)
    
    # STEP 5: Export
    logger.info("=" * 60)
    logger.info("📤 STEP 5: Export")
    logger.info("=" * 60)
    
    output_path = export_to_excel(companies)
    
    if not args.skip_db:
        export_to_database(companies)
    else:
        logger.info("   ⏭ Skipping database push")
    
    # Summary
    total_titles = sum(c.get('num_titles', 0) for c in companies)
    total_contacts = sum(c.get('num_contacts', 0) for c in companies)
    tmdb_enriched = sum(1 for c in companies if c.get('data_source') == 'tmdb')
    google_enriched = sum(1 for c in companies if c.get('data_source') == 'google_search')
    
    logger.info("=" * 70)
    logger.info("✅ ACCOUNTS PIPELINE COMPLETE")
    logger.info("=" * 70)
    logger.info(f"   📊 Summary:")
    logger.info(f"      Total companies: {len(companies)}")
    logger.info(f"      Total titles: {total_titles:,}")
    logger.info(f"      Total contacts: {total_contacts}")
    logger.info(f"   🔍 Data Sources:")
    logger.info(f"      TMDb enriched: {tmdb_enriched}")
    logger.info(f"      Google Search enriched: {google_enriched}")
    logger.info(f"      From catalog notes: {len(companies) - tmdb_enriched - google_enriched}")
    logger.info(f"   📁 Output: {output_path}")
    
    # Top 10 by titles
    logger.info("🏆 Top 10 Companies by Catalog Size:")
    sorted_companies = sorted(companies, key=lambda x: x.get('num_titles', 0), reverse=True)
    for c in sorted_companies[:10]:
        logger.info(f"   {c['num_titles']:>6,} | {c['name']:<35} | {c['num_contacts']} contacts")


if __name__ == "__main__":
    main()
