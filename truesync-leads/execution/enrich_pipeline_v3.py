#!/usr/bin/env python3
"""
TrueSync Lead Enrichment Pipeline v3
=====================================

A high-performance lead enrichment pipeline that uses BATCH processing
for all API calls to maximize efficiency and minimize run time.

KEY FEATURES:
- Batch Google Search: 20 queries per API call (vs 1 at a time)
- Batch leads-finder: All company domains in single call
- Professional Excel output with styling
- Database integration (optional)
- Retry logic with exponential backoff
- API result caching
- Input validation
- Structured logging

PIPELINE FLOW:
1. Load Sales Nav leads from JSON
2. PART 1: Find LinkedIn URLs via batched Google search
3. PART 2: Enrich with verified emails (batched per domain group)
4. PART 3: Discover additional leads at target companies
5. PART 4: Export to styled Excel + push to database

PERFORMANCE:
- 55 leads: ~3-5 minutes (vs ~20 minutes with single queries)
- 100 leads: ~5-8 minutes
"""
import os
import sys
import json
import time
import argparse
import uuid
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import centralized utilities and data
from data.companies import get_company_domains, get_domain_for_company, generate_email
from data.markets import detect_market, detect_markets_from_leads, get_location_filter_for_market
from utils import (
    setup_logging, get_logger, retry_with_backoff, APICache,
    validate_leads, validate_json_file, name_similarity, ProgressTracker
)

# ICP role keywords for filtering
ICP_ROLES_HIGH = [
    'acquisitions', 'distribution', 'licensing', 'content sales',
    'content partnerships', 'programming', 'content strategy'
]

ICP_ROLES_MEDIUM = [
    'head of content', 'vp content', 'svp content', 'evp content',
    'director content', 'global head'
]

# Job titles for lead discovery
DISCOVERY_JOB_TITLES = [
    # Acquisitions
    "Head of Acquisitions",
    "VP Acquisitions", 
    "Director of Acquisitions",
    # Distribution
    "Head of Distribution",
    "VP Distribution",
    # Licensing & Sales
    "Head of Licensing",
    "Head of Content Sales",
    "VP Content Partnerships",
    # Programming & Content
    "Head of Programming",
    "Head of Content",
    "SVP Content",
    "EVP Content",
    # Consumer Insights / Data Science (Validation Partners)
    "Head of Consumer Insights",
    "VP Data Science",
    "Head of Content Analytics",
    "VP Consumer Research",
    "Director of Audience Insights",
    # Studio Business Owners
    "President of Studio",
    "Studio Head",
    "Head of Label",
    "EVP Franchise Development",
    # International Originals Leadership
    "Head of International Originals",
    "VP Local Originals",
    "Head of Local Content",
    # Universal-specific targets
    "CFO",
    "SVP Commercial Strategy",
    "SVP Production",
    "SVP Programming",
    "SVP Operations",
]


def score_lead(lead: dict) -> int:
    """
    Score a lead based on ICP match.
    
    Uses the same logic as the centralized LeadScorer but works with dicts.
    Scoring factors:
    - Company tier (25 points max)
    - Role relevance (40 points max)
    - Seniority (25 points max)
    - Scope/Geography (15 points max)
    
    Total max: 100 points
    """
    score = 0
    title = (lead.get('title') or '').lower()
    company = (lead.get('company') or '').lower()
    
    # Company tier scoring (major entertainment/streaming companies)
    tier1_companies = ['netflix', 'amazon', 'prime video', 'warner bros', 'disney', 'paramount', 'sony pictures']
    tier2_companies = ['lionsgate', 'studiocanal', 'canal+', 'sky', 'bbc', 'itv', 'channel 4', 'mgm', 'hbo', 'max']
    tier3_companies = ['pluto tv', 'tubi', 'roku', 'freevee', 'peacock', 'gaumont', 'beta film', 'all3media']
    
    if any(c in company for c in tier1_companies):
        score += 25
    elif any(c in company for c in tier2_companies):
        score += 20
    elif any(c in company for c in tier3_companies):
        score += 15
    else:
        score += 10  # Unknown but still media/entertainment
    
    # Role relevance - HIGH priority roles
    for keyword in ICP_ROLES_HIGH:
        if keyword in title:
            score += 40
            break
    else:
        # MEDIUM priority roles
        for keyword in ICP_ROLES_MEDIUM:
            if keyword in title:
                score += 25
                break
    
    # Seniority scoring
    if any(x in title for x in ['evp', 'executive vice president', 'chief', 'president']):
        score += 25
    elif any(x in title for x in ['svp', 'senior vice president']):
        score += 20
    elif any(x in title for x in ['vp', 'vice president']):
        score += 15
    elif any(x in title for x in ['director', 'head of', 'global head', 'group head']):
        score += 12
    elif any(x in title for x in ['senior manager', 'senior director']):
        score += 10
    
    # Scope multiplier
    if any(x in title for x in ['global', 'worldwide', 'international']):
        score += 15
    elif any(x in title for x in ['group', 'executive']):
        score += 10
    elif any(x in title for x in ['regional', 'emea', 'apac', 'latam', 'americas']):
        score += 8
    
    return min(score, 100)


# ============================================================================
# PART 1: Find LinkedIn URLs using Google Search (anchor/linkedin-people-finder)
# ============================================================================

# Initialize cache for API results
_api_cache = None

def get_api_cache() -> APICache:
    """Get or create the API cache instance."""
    global _api_cache
    if _api_cache is None:
        _api_cache = APICache(cache_dir='.cache', ttl_hours=24)
    return _api_cache


def find_linkedin_urls(leads: list, use_cache: bool = True) -> list:
    """
    Find LinkedIn URLs for leads using Google Search.
    Searches for: "Name" "Company" site:linkedin.com/in
    
    Args:
        leads: List of lead dictionaries
        use_cache: Whether to use cached results
        
    Returns:
        Leads with LinkedIn URLs populated
    """
    logger = get_logger()
    
    token = os.getenv('APIFY_TOKEN')
    if not token:
        logger.error("❌ APIFY_TOKEN not found")
        return leads
    
    from apify_client import ApifyClient
    client = ApifyClient(token)
    
    logger.info("=" * 60)
    logger.info("🔗 PART 1: Find LinkedIn URLs via Google Search")
    logger.info("=" * 60)
    logger.info(f"   Using: apify/google-search-scraper")
    logger.info(f"   Total leads: {len(leads)}")
    
    # Find leads without LinkedIn URLs
    leads_needing_urls = [l for l in leads if not l.get('linkedin_url')]
    
    if not leads_needing_urls:
        logger.info("   ✅ All leads already have LinkedIn URLs")
        return leads
    
    logger.info(f"   Leads needing LinkedIn URLs: {len(leads_needing_urls)}")
    
    # Prepare search queries
    people_to_search = [
        {
            "name": lead.get('name', ''),
            "company": lead.get('company', ''),
        }
        for lead in leads_needing_urls
    ]
    
    # Get URLs via Google Search
    logger.info("   🔍 Searching Google for LinkedIn profiles...")
    name_to_url = google_search_linkedin_urls_batch(people_to_search, client, use_cache)
    
    logger.info(f"   Found {len(name_to_url)} LinkedIn URLs via Google Search")
    
    # Match URLs back to leads
    urls_matched = 0
    for lead in leads:
        if not lead.get('linkedin_url'):
            name = lead.get('name', '')
            # Try exact match first
            if name in name_to_url:
                lead['linkedin_url'] = name_to_url[name]
                urls_matched += 1
            else:
                # Try fuzzy match
                for search_name, url in name_to_url.items():
                    if name_similarity(name, search_name) > 0.85:
                        lead['linkedin_url'] = url
                        urls_matched += 1
                        break
    
    logger.info(f"   ✅ LinkedIn URLs matched: {urls_matched}/{len(leads_needing_urls)}")
    
    return leads


@retry_with_backoff(max_retries=3, base_delay=5.0, exceptions=(Exception,))
def _run_google_search_batch(client, queries_string: str) -> list:
    """Run a single batch of Google searches with retry logic."""
    actor_id = "apify/google-search-scraper"
    run = client.actor(actor_id).call(
        run_input={
            "queries": queries_string,
            "maxPagesPerQuery": 1,
            "resultsPerPage": 3,
        },
        timeout_secs=300,
        memory_mbytes=1024
    )
    return list(client.dataset(run["defaultDatasetId"]).iterate_items())


def google_search_linkedin_urls_batch(people: list, client, use_cache: bool = True) -> dict:
    """
    Use Google Search to find LinkedIn URLs for multiple people.
    Batches multiple queries into single API calls for efficiency.
    
    Args:
        people: List of dicts with 'name' and 'company' keys
        client: Apify client instance
        use_cache: Whether to use cached results
        
    Returns:
        Dict mapping names to LinkedIn URLs
    """
    logger = get_logger()
    cache = get_api_cache() if use_cache else None
    name_to_url = {}
    
    # Build all queries
    queries = []
    query_to_name = {}
    
    for person in people:
        name = person.get("name", "")
        company = person.get("company", "")
        
        if not name:
            continue
        
        query = f'"{name}" "{company}" site:linkedin.com/in'
        queries.append(query)
        query_to_name[query] = name
    
    if not queries:
        return name_to_url
    
    # Process in batches of 20 queries per API call
    batch_size = 20
    total_batches = (len(queries) + batch_size - 1) // batch_size
    
    for i in range(0, len(queries), batch_size):
        batch_queries = queries[i:i+batch_size]
        batch_num = (i // batch_size) + 1
        
        logger.info(f"   📍 Batch {batch_num}/{total_batches} ({len(batch_queries)} queries)...")
        
        # Check cache first
        queries_string = "\n".join(batch_queries)
        cache_key = f"google_search_{hash(queries_string)}"
        
        search_results = None
        if cache:
            search_results = cache.get(cache_key)
            if search_results:
                logger.info(f"      📦 Using cached results")
        
        if search_results is None:
            try:
                search_results = _run_google_search_batch(client, queries_string)
                
                # Cache successful results
                if cache and search_results:
                    cache.set(cache_key, search_results)
                    
            except Exception as e:
                logger.error(f"      ⚠ Batch error: {str(e)[:60]}")
                continue
        
        # Process results
        for sr in search_results:
            search_query = sr.get('searchQuery', {})
            query_term = search_query.get('term', '')
            organic = sr.get('organicResults', [])
            
            # Find the name this query was for
            name = query_to_name.get(query_term, '')
            if not name:
                for q, n in query_to_name.items():
                    if n.lower() in query_term.lower():
                        name = n
                        break
            
            if not name:
                continue
            
            # Find LinkedIn URL in results
            for result in organic:
                url = result.get('url', '')
                if 'linkedin.com/in/' in url:
                    name_to_url[name] = url
                    logger.debug(f"      ✓ {name}: {url[:50]}...")
                    break
        
        found_count = len([q for q in batch_queries if query_to_name.get(q, '') in name_to_url])
        logger.info(f"      Found {found_count} URLs in this batch")
    
    return name_to_url


# ============================================================================
# PART 2: Enrich with Emails using code_crafter/leads-finder
# ============================================================================

@retry_with_backoff(max_retries=3, base_delay=5.0, exceptions=(Exception,))
def _run_leads_finder_batch(client, run_input: dict) -> list:
    """Run leads-finder actor with retry logic."""
    actor_id = "code_crafter/leads-finder"
    run = client.actor(actor_id).call(
        run_input=run_input,
        timeout_secs=300,
        memory_mbytes=1024
    )
    return list(client.dataset(run["defaultDatasetId"]).iterate_items())


def enrich_leads_from_apify(leads: list, location: str = "", use_cache: bool = True) -> list:
    """
    Search for leads via Apify to get verified emails.
    Uses code_crafter/leads-finder with BATCHED domain queries.
    
    Args:
        leads: List of lead dictionaries
        location: Optional location filter
        use_cache: Whether to use cached results
        
    Returns:
        Tuple of (enriched leads, apify results for discovery)
    """
    logger = get_logger()
    cache = get_api_cache() if use_cache else None
    
    token = os.getenv('APIFY_TOKEN')
    if not token:
        logger.error("❌ APIFY_TOKEN not found")
        return leads, []
    
    from apify_client import ApifyClient
    client = ApifyClient(token)
    
    actor_id = "code_crafter/leads-finder"
    
    logger.info("=" * 60)
    logger.info("📧 PART 2: Enrich with Verified Emails (Batched)")
    logger.info("=" * 60)
    logger.info(f"   Actor: {actor_id}")
    
    # Get unique companies from leads and find their domains using centralized function
    unique_companies = set(l.get('company', '') for l in leads if l.get('company'))
    company_domains = {}
    for company in unique_companies:
        domain = get_domain_for_company(company)
        if domain:
            company_domains[company] = domain
    
    # Get unique domains
    domains = list(set(company_domains.values()))
    
    if not domains:
        logger.warning("   ⚠ No company domains found - skipping email enrichment via leads-finder")
        logger.info("   💡 Will attempt email lookup via LinkedIn profiles instead")
        return leads, []
    
    logger.info(f"   Found {len(domains)} unique company domains from {len(unique_companies)} companies")
    logger.info(f"   Domains: {', '.join(domains[:10])}{'...' if len(domains) > 10 else ''}")
    
    all_apify_results = []
    
    # Use centralized market detection
    if not location:
        detected_markets = detect_markets_from_leads(leads)
        locations = []
        for market in detected_markets:
            locations.extend(get_location_filter_for_market(market))
        locations = list(set(locations))  # Remove duplicates
        logger.info(f"   Detected markets: {', '.join(locations) if locations else 'Global (no filter)'}")
    else:
        locations = [location]
    
    # BATCH: Search all domains in one API call
    logger.info(f"   📍 Batch search: {', '.join(domains[:5])}{'...' if len(domains) > 5 else ''}")
    
    run_input = {
        "contact_company_domain": domains,
        "contact_job_title": DISCOVERY_JOB_TITLES,
        "seniority_level": ["head", "director", "vp", "c_suite", "manager"],
        "fetch_count": 300,
    }
    
    if locations:
        run_input["contact_location"] = locations
    
    # Check cache first
    cache_key = f"leads_finder_{hash(str(sorted(domains)) + str(sorted(locations)))}"
    results = None
    
    if cache:
        results = cache.get(cache_key)
        if results:
            logger.info(f"      📦 Using cached results ({len(results)} profiles)")
    
    if results is None:
        try:
            results = _run_leads_finder_batch(client, run_input)
            logger.info(f"      Found {len(results)} profiles in batch search")
            
            # Cache successful results
            if cache and results:
                cache.set(cache_key, results)
                
        except Exception as e:
            logger.error(f"      ⚠ Batch error: {str(e)[:60]}")
            logger.info(f"      🔄 Falling back to per-domain searches...")
            results = []
            
            # Fallback: per-domain searches if batch fails
            for domain in domains[:10]:
                try:
                    fallback_input = {
                        "contact_company_domain": [domain],
                        "contact_job_title": DISCOVERY_JOB_TITLES,
                        "seniority_level": ["head", "director", "vp", "c_suite", "manager"],
                        "fetch_count": 100,
                    }
                    if locations:
                        fallback_input["contact_location"] = locations
                    
                    domain_results = _run_leads_finder_batch(client, fallback_input)
                    logger.info(f"      {domain}: {len(domain_results)} profiles")
                    results.extend(domain_results)
                    
                except Exception as e2:
                    logger.warning(f"      ⚠ {domain}: {str(e2)[:40]}")
                    continue
    
    all_apify_results.extend(results)
    logger.info(f"   Total Apify profiles found: {len(all_apify_results)}")
    
    # Match Apify results back to leads for emails (and LinkedIn URLs if still missing)
    email_found = 0
    linkedin_found = 0
    
    for lead in leads:
        lead_name = lead.get('name', '').lower().strip()
        
        best_match = None
        best_score = 0
        
        for result in all_apify_results:
            result_name = f"{result.get('first_name', '')} {result.get('last_name', '')}".strip()
            if not result_name:
                result_name = result.get('full_name', result.get('name', ''))
            
            similarity = name_similarity(lead_name, result_name)
            if similarity > best_score and similarity > 0.75:
                best_score = similarity
                best_match = result
        
        if best_match:
            # Get verified email
            email = best_match.get('email', '')
            if email and not lead.get('email'):
                lead['email'] = email
                lead['email_verified'] = True
                email_found += 1
            
            # Get LinkedIn URL if still missing
            linkedin = best_match.get('linkedin') or best_match.get('linkedin_url', '')
            if linkedin and not lead.get('linkedin_url'):
                lead['linkedin_url'] = linkedin
                linkedin_found += 1
    
    logger.info(f"   ✅ Matched Results:")
    logger.info(f"      Verified emails: {email_found}/{len(leads)}")
    logger.info(f"      LinkedIn URLs (additional): {linkedin_found}")
    
    # Apply pattern-based emails for remaining leads using centralized function
    pattern_count = 0
    for lead in leads:
        if not lead.get('email'):
            domain = get_domain_for_company(lead.get('company', ''))
            if domain:
                name_parts = lead.get('name', '').replace('.', ' ').split()
                if len(name_parts) >= 2:
                    first_name = name_parts[0]
                    last_name = name_parts[-1]
                    # Use centralized email generation with company-specific patterns
                    email = generate_email(first_name, last_name, domain)
                    if email:
                        lead['email'] = email
                        lead['email_verified'] = False
                        pattern_count += 1
    
    if pattern_count > 0:
        logger.info(f"      Pattern-based emails added: {pattern_count}")
    
    return leads, all_apify_results


# ============================================================================
# PART 3: Discover Additional Leads
# ============================================================================

def discover_additional_leads(existing_leads: list, apify_results: list, location: str = "united kingdom") -> list:
    """
    Find NEW leads from Apify results that weren't in Sales Nav export.
    Only includes leads from target companies defined in data/companies.py.
    
    Args:
        existing_leads: List of existing lead dictionaries
        apify_results: Results from Apify leads-finder
        location: Default location for new leads
        
    Returns:
        List of newly discovered leads
    """
    logger = get_logger()
    
    logger.info("=" * 60)
    logger.info("🔍 PART 3: Extract Additional Leads from Apify Results")
    logger.info("=" * 60)
    
    # Create set of existing names for deduplication
    existing_names = {lead.get('name', '').lower().strip() for lead in existing_leads}
    
    new_leads = []
    filtered_non_target = 0
    
    for result in apify_results:
        name = f"{result.get('first_name', '')} {result.get('last_name', '')}".strip()
        if not name:
            name = result.get('full_name', result.get('name', ''))
        
        if not name or name.lower().strip() in existing_names:
            continue
        
        # Filter to target companies only - skip leads from non-target companies
        company_name = result.get('company_name', '')
        if not get_domain_for_company(company_name):
            filtered_non_target += 1
            continue
        
        # Check if this is a relevant lead
        title = result.get('job_title', result.get('title', result.get('headline', '')))
        if not title:
            continue
        
        # Detect market from location
        result_location = result.get('country', location)
        detected_market = detect_market(result_location)
        
        # Create lead
        new_lead = {
            'name': name,
            'title': title,
            'company': company_name,
            'email': result.get('email', ''),
            'email_verified': bool(result.get('email')),
            'linkedin_url': result.get('linkedin') or result.get('linkedin_url', ''),
            'location': result_location,
            'detected_market': detected_market,
        }
        new_lead['score'] = score_lead(new_lead)
        
        # Only include if ICP score is decent
        if new_lead['score'] >= 50:
            new_leads.append(new_lead)
            existing_names.add(name.lower().strip())
    
    logger.info(f"   ✅ Discovered {len(new_leads)} NEW leads from Apify")
    verified = sum(1 for l in new_leads if l.get('email_verified'))
    logger.info(f"      With verified emails: {verified}")
    if filtered_non_target > 0:
        logger.info(f"      Filtered out {filtered_non_target} leads from non-target companies")
    
    return new_leads


# ============================================================================
# PART 4: Export with Professional Excel Styling
# ============================================================================

def export_to_styled_excel(leads: list, market: str = "UK") -> str:
    """
    Export leads to professionally styled Excel workbook.
    Matches the style from the original output.
    
    Args:
        leads: List of lead dictionaries
        market: Default market label
        
    Returns:
        Path to exported Excel file
    """
    logger = get_logger()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f"output/truesync_leads_v3_{timestamp}.xlsx"
    
    Path("output").mkdir(exist_ok=True)
    
    # Prepare data (without Email Source and Source columns)
    df_data = []
    for lead in leads:
        df_data.append({
            'Name': lead.get('name', ''),
            'Title': lead.get('title', ''),
            'Company': lead.get('company', ''),
            'LinkedIn URL': lead.get('linkedin_url', ''),
            'Email': lead.get('email', ''),
            'Location': lead.get('location', ''),
            'Priority Score': lead.get('score', 0),
            'Market': market,
        })
    
    df = pd.DataFrame(df_data)
    df = df.sort_values('Priority Score', ascending=False)
    
    # Create styled workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Lead List"
    
    # Styling definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    score_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # Green
    score_medium = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    score_low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Red
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    column_widths = {
        'A': 25,  # Name
        'B': 45,  # Title
        'C': 20,  # Company
        'D': 45,  # LinkedIn URL
        'E': 35,  # Email
        'F': 25,  # Location
        'G': 15,  # Priority Score
        'H': 10,  # Market
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            
            # Color-code Priority Score column
            if col_idx == 7:  # Priority Score column
                try:
                    score = int(value or 0)
                    if score >= 75:
                        cell.fill = score_high
                    elif score >= 50:
                        cell.fill = score_medium
                    else:
                        cell.fill = score_low
                except (ValueError, TypeError):
                    pass
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions
    
    # Create market-specific sheet
    uk_leads = [l for l in leads if 'UK' in (l.get('market') or market).upper()]
    if uk_leads:
        ws_uk = wb.create_sheet("UK Leads")
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws_uk.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write UK data
        uk_df = df[df['Market'] == market].sort_values('Priority Score', ascending=False)
        for row_idx, row in enumerate(uk_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_uk.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border
                if col_idx == 7:
                    try:
                        score = int(value or 0)
                        if score >= 75:
                            cell.fill = score_high
                        elif score >= 50:
                            cell.fill = score_medium
                        else:
                            cell.fill = score_low
                    except:
                        pass
        
        # Set column widths
        for col, width in column_widths.items():
            ws_uk.column_dimensions[col].width = width
        
        ws_uk.freeze_panes = 'A2'
    
    # Save
    wb.save(output_path)
    
    # Also save CSV
    csv_path = output_path.replace('.xlsx', '.csv')
    df.to_csv(csv_path, index=False)
    
    logger.info(f"📊 Exported to: {output_path}")
    logger.info(f"   CSV: {csv_path}")
    logger.info(f"   Total leads: {len(leads)}")
    
    return output_path


def push_to_database(leads: list, market: str = 'UK') -> int:
    """
    Push leads to database.
    
    Args:
        leads: List of lead dictionaries
        market: Default market label
        
    Returns:
        Number of leads inserted
    """
    logger = get_logger()
    
    try:
        from db.connection import get_engine
        from db.models import Lead, Company
        from sqlalchemy.orm import Session
        from sqlalchemy import select
    except ImportError as e:
        logger.warning(f"⚠ Database modules not available: {e}")
        return 0
    
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        logger.warning("⚠ DATABASE_URL not found")
        return 0
    
    logger.info(f"💾 Pushing {len(leads)} leads to database...")
    
    try:
        engine = get_engine()
        inserted = 0
        updated = 0
        
        with Session(engine) as session:
            company_ids = {}
            
            for lead in leads:
                company_name = lead.get('company', 'Unknown')
                
                if company_name not in company_ids:
                    stmt = select(Company).where(Company.name == company_name)
                    company = session.execute(stmt).scalar_one_or_none()
                    
                    if company:
                        company_ids[company_name] = company.id
                    else:
                        new_company = Company(
                            id=str(uuid.uuid4()),
                            name=company_name,
                            type='platform' if 'sky' in company_name.lower() else 'producer',
                            market=market
                        )
                        session.add(new_company)
                        session.flush()
                        company_ids[company_name] = new_company.id
                
                # Check for existing lead by name or LinkedIn URL
                linkedin_url = lead.get('linkedin_url') or None
                stmt = select(Lead).where(Lead.name == lead.get('name'))
                existing = session.execute(stmt).scalar_one_or_none()
                
                if existing:
                    if lead.get('email') and not existing.email:
                        existing.email = lead['email']
                        updated += 1
                    if linkedin_url and not existing.linkedin_url:
                        existing.linkedin_url = linkedin_url
                else:
                    new_lead = Lead(
                        id=str(uuid.uuid4()),
                        name=lead.get('name'),
                        title=lead.get('title'),
                        company_id=company_ids.get(company_name),
                        email=lead.get('email'),
                        market=market,
                        priority_score=lead.get('score', 0),
                        linkedin_url=linkedin_url if linkedin_url else None,
                    )
                    session.add(new_lead)
                    inserted += 1
            
            session.commit()
        
        logger.info(f"   ✅ Inserted: {inserted}, Updated: {updated}")
        return inserted
        
    except Exception as e:
        logger.error(f"   ❌ Database error: {e}")
        return 0


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="TrueSync Lead Pipeline v3 - Enhanced with caching, retry, and logging")
    parser.add_argument("--input", default=".tmp/uk_leads_raw.json", help="Input JSON file")
    parser.add_argument("--market", default="Global", help="Market name")
    parser.add_argument("--location", default="", help="Location filter (optional)")
    parser.add_argument("--skip-enrich", action="store_true", help="Skip enrichment")
    parser.add_argument("--skip-discovery", action="store_true", help="Skip lead discovery")
    parser.add_argument("--skip-db", action="store_true", help="Skip database push")
    parser.add_argument("--skip-cache", action="store_true", help="Skip API result caching")
    parser.add_argument("--filter-companies", action="store_true", help="Only include target companies")
    parser.add_argument("--clear-cache", action="store_true", help="Clear API cache before running")
    
    args = parser.parse_args()
    
    # Set up logging
    logger = setup_logging(output_dir='output', log_to_file=True)
    use_cache = not args.skip_cache
    
    logger.info("=" * 70)
    logger.info("🚀 TRUESYNC LEAD PIPELINE v3")
    logger.info("   - Google Search for LinkedIn URLs (name + company)")
    logger.info("   - Verified emails via Apify")
    logger.info("   - Professional Excel styling")
    logger.info("   - Multi-market support (USA, UK, France, Spain, Korea, Germany)")
    logger.info("   - Retry logic with exponential backoff")
    logger.info("   - API result caching (24h TTL)")
    logger.info("=" * 70)
    
    # Clear cache if requested
    if args.clear_cache:
        cache = get_api_cache()
        cache.clear()
        logger.info("   Cache cleared")
    
    # Load and validate leads
    input_path = Path(__file__).parent.parent / args.input
    logger.info(f"📂 Loading leads from: {input_path}")
    
    try:
        leads = validate_json_file(str(input_path))
    except ValueError as e:
        logger.error(f"❌ {e}")
        return
    
    logger.info(f"   Loaded and validated {len(leads)} leads")
    
    # Optional: Filter to target companies only
    if args.filter_companies:
        original_count = len(leads)
        leads = [l for l in leads if get_domain_for_company(l.get('company', ''))]
        logger.info(f"   Filtered to {len(leads)} leads in target companies (from {original_count})")
    
    # Score leads and assign market
    for lead in leads:
        lead['score'] = score_lead(lead)
        lead['market'] = args.market
        # Also detect market from location if available
        if lead.get('location'):
            lead['detected_market'] = detect_market(lead['location'])
    
    # PART 1 & 2: Find LinkedIn URLs and enrich emails
    apify_results = []
    if not args.skip_enrich:
        leads = find_linkedin_urls(leads, use_cache=use_cache)
        leads, apify_results = enrich_leads_from_apify(leads, args.location, use_cache=use_cache)
    else:
        logger.info("⏭ Skipping enrichment")
        # Add pattern emails for leads without email
        for lead in leads:
            if not lead.get('email'):
                domain = get_domain_for_company(lead.get('company', ''))
                if domain:
                    name_parts = lead.get('name', '').replace('.', ' ').split()
                    if len(name_parts) >= 2:
                        email = generate_email(name_parts[0], name_parts[-1], domain)
                        if email:
                            lead['email'] = email
                            lead['email_verified'] = False
    
    # PART 3: Discover additional leads from Apify results
    new_leads = []
    if not args.skip_discovery and apify_results:
        new_leads = discover_additional_leads(leads, apify_results, args.location)
        for lead in new_leads:
            lead['market'] = args.market
        leads.extend(new_leads)
    elif args.skip_discovery:
        logger.info("⏭ Skipping lead discovery")
    
    # Sort all leads by score
    leads.sort(key=lambda x: x.get('score', 0), reverse=True)
    
    # PART 4: Export
    logger.info("=" * 60)
    logger.info("📤 PART 4: Export & Database")
    logger.info("=" * 60)
    
    output_path = export_to_styled_excel(leads, args.market)
    
    if not args.skip_db:
        push_to_database(leads, args.market)
    else:
        logger.info("⏭ Skipping database push")
    
    # Summary
    verified = sum(1 for l in leads if l.get('email_verified'))
    pattern = sum(1 for l in leads if not l.get('email_verified') and l.get('email'))
    with_linkedin = sum(1 for l in leads if l.get('linkedin_url'))
    
    logger.info("=" * 70)
    logger.info("✅ PIPELINE COMPLETE")
    logger.info("=" * 70)
    logger.info(f"   📊 Lead Summary:")
    logger.info(f"      Total leads: {len(leads)}")
    logger.info(f"      Sales Navigator leads: {len(leads) - len(new_leads)}")
    logger.info(f"      Newly discovered leads: {len(new_leads)}")
    logger.info(f"   🔗 LinkedIn URLs: {with_linkedin}/{len(leads)}")
    logger.info(f"   📧 Email Summary:")
    logger.info(f"      Verified (Apify): {verified}")
    logger.info(f"      Pattern-based: {pattern}")
    logger.info(f"   Average ICP score: {sum(l.get('score', 0) for l in leads) / max(len(leads), 1):.1f}")
    logger.info(f"   Output: {output_path}")
    
    # Top leads
    logger.info("🏆 Top 15 Leads:")
    for lead in leads[:15]:
        email = lead.get('email', 'N/A')[:30]
        icon = "✅" if lead.get('email_verified') else "📧"
        linkedin = "🔗" if lead.get('linkedin_url') else "  "
        logger.info(f"   {lead.get('score', 0):>3} | {lead['name']:<25} | {email:<30} {icon} {linkedin}")


if __name__ == "__main__":
    main()
