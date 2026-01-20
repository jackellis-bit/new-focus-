#!/usr/bin/env python3
"""
TrueSync Lead Generation Pipeline
==================================

Full pipeline for finding and enriching LinkedIn leads:
1. Google Search → Find LinkedIn URLs for ICP roles at target companies
2. LinkedIn Profile Scraper → Enrich with emails
3. TMDb → Add catalog data (shows, ratings)
4. Export to Excel with scoring

Usage:
    python run_batch_test.py                    # Test with 1 company
    python run_batch_test.py --tier1            # Run Tier 1 markets
    python run_batch_test.py --tier2            # Run Tier 2 markets
    python run_batch_test.py --all              # Run all companies
    python run_batch_test.py --avod             # Run AVOD platforms only
"""

import os
import sys
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import yaml
from dotenv import load_dotenv
load_dotenv()

# Force real Apify
os.environ['USE_MOCK_DATA'] = 'false'

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

try:
    from apify_client import ApifyClient
except ImportError:
    print("Error: apify-client not installed. Run: pip install apify-client")
    sys.exit(1)

# Import company data
from data.companies import (
    COMPANIES, get_tier1_companies, get_tier2_companies, 
    get_avod_companies, get_all_companies
)

# Import TMDb for catalog enrichment
from scrapers.catalog import TMDbClient


def load_config() -> dict:
    """Load configuration from config.yaml."""
    config_path = Path(__file__).parent / 'config.yaml'
    with open(config_path, 'r') as f:
        return yaml.safe_load(f)


def get_target_roles(config: dict) -> List[str]:
    """Get all target ICP roles from config."""
    roles = config.get('discovery', {}).get('target_roles', {})
    
    all_roles = []
    # Primary ICPs first (higher priority)
    for category in ['primary_acquisitions', 'primary_partnerships', 'primary_distribution']:
        all_roles.extend(roles.get(category, []))
    
    # Secondary ICPs
    for category in ['secondary_strategy', 'secondary_avod']:
        all_roles.extend(roles.get(category, []))
    
    return all_roles


def get_exclude_keywords(config: dict) -> List[str]:
    """Get roles to exclude from config."""
    return config.get('discovery', {}).get('exclude_roles', [])


def is_valid_lead(title: str, exclude_keywords: List[str]) -> bool:
    """Check if a lead's title is valid (not in exclusion list)."""
    if not title:
        return True
    
    title_lower = title.lower()
    for keyword in exclude_keywords:
        if keyword.lower() in title_lower:
            return False
    return True


def categorize_icp(title: str, config: dict) -> str:
    """Categorize a lead into ICP category based on title keywords."""
    if not title:
        return "Unknown"
    
    title_lower = title.lower()
    
    # Primary ICPs - check for key role keywords (more flexible matching)
    
    # Acquisitions keywords
    acquisitions_keywords = ['acquisition', 'acquisitions', 'content buyer', 'content buying']
    if any(kw in title_lower for kw in acquisitions_keywords):
        return "Primary - Acquisitions"
    
    # Partnerships keywords
    partnerships_keywords = ['partnership', 'partnerships', 'content partner', 'international content']
    if any(kw in title_lower for kw in partnerships_keywords):
        return "Primary - Partnerships"
    
    # Distribution/Licensing keywords  
    distribution_keywords = ['distribution', 'licensing', 'international sales', 'global sales', 
                            'content sales', 'global distribution']
    if any(kw in title_lower for kw in distribution_keywords):
        return "Primary - Distribution"
    
    # Strategy keywords
    strategy_keywords = ['content strategy', 'international strategy', 'portfolio strategy']
    if any(kw in title_lower for kw in strategy_keywords):
        return "Secondary - Strategy"
    
    # Programming keywords (AVOD)
    programming_keywords = ['programming', 'content programming']
    if any(kw in title_lower for kw in programming_keywords):
        return "Secondary - AVOD"
    
    # Also check for seniority + relevant area
    senior_keywords = ['vp', 'vice president', 'svp', 'evp', 'head of', 'director', 'chief']
    relevant_areas = ['content', 'international', 'global', 'film', 'series', 'original']
    
    has_seniority = any(kw in title_lower for kw in senior_keywords)
    has_relevant_area = any(kw in title_lower for kw in relevant_areas)
    
    if has_seniority and has_relevant_area:
        # Check config roles for more specific categorization
        roles = config.get('discovery', {}).get('target_roles', {})
        
        for role in roles.get('primary_acquisitions', []):
            if role.lower() in title_lower:
                return "Primary - Acquisitions"
        
        for role in roles.get('primary_partnerships', []):
            if role.lower() in title_lower:
                return "Primary - Partnerships"
        
        for role in roles.get('primary_distribution', []):
            if role.lower() in title_lower:
                return "Primary - Distribution"
        
        return "Secondary - Other"
    
    return "Other"


class LinkedInLeadGenerator:
    """Full pipeline for finding and enriching LinkedIn leads."""
    
    def __init__(self, config: dict):
        self.api_token = os.getenv('APIFY_TOKEN')
        if not self.api_token:
            raise ValueError("APIFY_TOKEN not set!")
        
        self.client = ApifyClient(self.api_token)
        self.config = config
        self.exclude_keywords = get_exclude_keywords(config)
        self.leads = []
        
        # Initialize TMDb client for catalog enrichment
        self.tmdb = TMDbClient()
        self.catalog_cache = {}  # Cache catalog data per company
    
    def _clean_company_name(self, company: str) -> str:
        """Clean company name for search (remove market suffixes)."""
        # Remove common market/region suffixes for broader search
        suffixes_to_remove = [' US', ' UK', ' Spain', ' Korea', ' Germany', ' France', ' DE']
        clean_name = company
        for suffix in suffixes_to_remove:
            if clean_name.endswith(suffix):
                clean_name = clean_name[:-len(suffix)]
        return clean_name
    
    def search_google_for_profiles(
        self,
        company: str,
        roles: List[str],
        max_results: int = 20
    ) -> List[Dict]:
        """Use Google to find LinkedIn profile URLs."""
        # Clean company name for broader search results
        search_company = self._clean_company_name(company)
        
        # Build search query with top roles
        role_query = ' OR '.join([f'"{role}"' for role in roles[:4]])
        query = f'site:linkedin.com/in "{search_company}" ({role_query})'
        
        print(f"  Searching: {search_company} + ICP roles...")
        
        run_input = {
            'queries': query,
            'maxPagesPerQuery': 2,
            'resultsPerPage': 20
        }
        
        try:
            run = self.client.actor('apify/google-search-scraper').call(
                run_input=run_input,
                timeout_secs=120
            )
            
            dataset_id = run.get('defaultDatasetId')
            if not dataset_id:
                return []
            
            linkedin_profiles = []
            for item in self.client.dataset(dataset_id).iterate_items():
                for result in item.get('organicResults', []):
                    url = result.get('url', '')
                    title = result.get('title', '')
                    
                    if 'linkedin.com/in/' in url and url not in [p['url'] for p in linkedin_profiles]:
                        linkedin_profiles.append({
                            'url': url,
                            'title': title,
                            'company': company
                        })
            
            print(f"  Found {len(linkedin_profiles)} LinkedIn profiles")
            return linkedin_profiles[:max_results]
            
        except Exception as e:
            print(f"  Error searching Google: {e}")
            return []
    
    def enrich_profiles(self, profiles: List[Dict]) -> List[Dict]:
        """Enrich LinkedIn profiles with full data and emails."""
        if not profiles:
            return []
        
        urls = [p['url'] for p in profiles]
        print(f"  Enriching {len(urls)} profiles with emails...")
        
        run_input = {
            'profileUrls': urls,
            'includeEmail': True
        }
        
        try:
            run = self.client.actor('dev_fusion/linkedin-profile-scraper').call(
                run_input=run_input,
                timeout_secs=300
            )
            
            dataset_id = run.get('defaultDatasetId')
            if not dataset_id:
                return profiles
            
            enriched = []
            for item in self.client.dataset(dataset_id).iterate_items():
                original = next(
                    (p for p in profiles if p['url'] in item.get('linkedinUrl', '')),
                    {}
                )
                
                title = item.get('headline', '')
                
                # Filter out excluded roles
                if not is_valid_lead(title, self.exclude_keywords):
                    print(f"    Excluding (non-ICP role): {title[:50]}...")
                    continue
                
                enriched.append({
                    'name': item.get('fullName', ''),
                    'title': title,
                    'email': item.get('email', ''),
                    'linkedin_url': item.get('linkedinUrl', ''),
                    'location': item.get('location', ''),
                    'about': item.get('about', ''),
                    'company': original.get('company', ''),
                    'connection_count': item.get('connectionCount', '')
                })
            
            emails_found = sum(1 for e in enriched if e.get('email'))
            print(f"  Enriched {len(enriched)} profiles, {emails_found} emails found")
            return enriched
            
        except Exception as e:
            print(f"  Error enriching profiles: {e}")
            return profiles
    
    def get_catalog_data(self, company_name: str, market: str) -> Dict:
        """Get catalog data for a company (with caching)."""
        cache_key = f"{company_name}_{market}"
        
        if cache_key not in self.catalog_cache:
            print(f"  Fetching TMDb catalog for {company_name}...")
            try:
                self.catalog_cache[cache_key] = self.tmdb.get_top_shows_formatted(
                    company_name, market, max_titles=5
                )
            except Exception as e:
                print(f"  TMDb error: {e}")
                self.catalog_cache[cache_key] = {
                    'top_shows': '',
                    'show_details': [],
                    'total_catalog': 0
                }
        
        return self.catalog_cache[cache_key]
    
    def run_batch(
        self,
        companies: List[Dict],
        roles: List[str],
        leads_per_company: int = 10
    ) -> List[Dict]:
        """Run the full pipeline for a batch of companies."""
        all_leads = []
        
        for company in companies:
            print(f"\n{'='*60}")
            print(f"Processing: {company['name']} ({company.get('market', 'usa').upper()})")
            print('='*60)
            
            # Step 1: Find LinkedIn profiles via Google
            profiles = self.search_google_for_profiles(
                company=company['name'],
                roles=roles,
                max_results=leads_per_company
            )
            
            if not profiles:
                print(f"  No profiles found for {company['name']}")
                continue
            
            # Step 2: Enrich with emails
            enriched = self.enrich_profiles(profiles)
            
            # Step 3: Get catalog data from TMDb
            catalog_data = self.get_catalog_data(
                company['name'], 
                company.get('market', 'usa')
            )
            
            # Add company metadata and catalog
            for lead in enriched:
                lead['company'] = company['name']
                lead['market'] = company.get('market', 'usa')
                lead['company_type'] = company.get('type', 'Studio')
                lead['catalog_size'] = company.get('catalog_size', '')
                lead['catalog_notes'] = company.get('catalog_notes', '')
                lead['icp_category'] = categorize_icp(lead.get('title', ''), self.config)
                
                # Add TMDb catalog data
                lead['top_shows'] = catalog_data.get('top_shows', '')
                lead['tmdb_catalog_count'] = catalog_data.get('total_catalog', 0)
            
            all_leads.extend(enriched)
            print(f"  Total leads from {company['name']}: {len(enriched)}")
        
        self.leads = all_leads
        return all_leads


class ExcelExporter:
    """Export leads to formatted Excel with ICP categorization."""
    
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    
    SCORE_COLORS = {
        'high': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    }
    
    ICP_COLORS = {
        'Primary - Acquisitions': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'Primary - Partnerships': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
        'Primary - Distribution': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'Secondary - Strategy': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        'Secondary - AVOD': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    }
    
    def __init__(self, output_dir: str = 'output'):
        self.output_dir = output_dir
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    def calculate_score(self, lead: Dict) -> int:
        """Calculate priority score for a lead based on ICP category."""
        score = 40  # Base score
        
        title = (lead.get('title') or '').lower()
        icp = lead.get('icp_category', '')
        
        # ICP category scoring
        if 'Primary - Acquisitions' in icp:
            score += 30  # Highest value
        elif 'Primary - Partnerships' in icp:
            score += 25
        elif 'Primary - Distribution' in icp:
            score += 25
        elif 'Secondary' in icp:
            score += 15
        
        # Seniority scoring
        if any(s in title for s in ['evp', 'svp', 'chief', 'head of']):
            score += 15
        elif any(s in title for s in ['vp', 'vice president', 'director']):
            score += 10
        
        # Email bonus
        if lead.get('email'):
            score += 15
        
        return min(score, 100)
    
    def export(self, leads: List[Dict], filename: str = None) -> str:
        """Export leads to Excel."""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"truesync_leads_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Calculate scores
        for lead in leads:
            lead['priority_score'] = self.calculate_score(lead)
        
        # Sort by score
        leads = sorted(leads, key=lambda x: x.get('priority_score', 0), reverse=True)
        
        wb = Workbook()
        
        # Master sheet
        ws = wb.active
        ws.title = "Master Lead List"
        
        headers = ['Name', 'Title', 'ICP Category', 'Company', 'Company Type', 
                   'LinkedIn URL', 'Email', 'Market', 'Priority Score', 
                   'Top Shows', 'Catalog Size', 'Location']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, lead in enumerate(leads, 2):
            ws.cell(row=row_idx, column=1, value=lead.get('name', ''))
            ws.cell(row=row_idx, column=2, value=lead.get('title', ''))
            
            icp_cell = ws.cell(row=row_idx, column=3, value=lead.get('icp_category', ''))
            icp = lead.get('icp_category', '')
            if icp in self.ICP_COLORS:
                icp_cell.fill = self.ICP_COLORS[icp]
                icp_cell.font = Font(color="FFFFFF", bold=True)
            
            ws.cell(row=row_idx, column=4, value=lead.get('company', ''))
            ws.cell(row=row_idx, column=5, value=lead.get('company_type', ''))
            ws.cell(row=row_idx, column=6, value=lead.get('linkedin_url', ''))
            ws.cell(row=row_idx, column=7, value=lead.get('email', ''))
            ws.cell(row=row_idx, column=8, value=lead.get('market', '').upper())
            
            score_cell = ws.cell(row=row_idx, column=9, value=lead.get('priority_score', 0))
            score = lead.get('priority_score', 0)
            if score >= 75:
                score_cell.fill = self.SCORE_COLORS['high']
            elif score >= 50:
                score_cell.fill = self.SCORE_COLORS['medium']
            else:
                score_cell.fill = self.SCORE_COLORS['low']
            
            ws.cell(row=row_idx, column=10, value=lead.get('top_shows', ''))
            ws.cell(row=row_idx, column=11, value=lead.get('catalog_size', ''))
            ws.cell(row=row_idx, column=12, value=lead.get('location', ''))
        
        widths = [25, 50, 22, 25, 15, 45, 35, 10, 15, 60, 30, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        
        # Primary ICPs sheet
        ws_primary = wb.create_sheet("Primary ICPs")
        primary_leads = [l for l in leads if 'Primary' in l.get('icp_category', '')]
        self._write_leads_to_sheet(ws_primary, primary_leads, headers, widths)
        
        # Secondary ICPs sheet
        ws_secondary = wb.create_sheet("Secondary ICPs")
        secondary_leads = [l for l in leads if 'Secondary' in l.get('icp_category', '')]
        self._write_leads_to_sheet(ws_secondary, secondary_leads, headers, widths)
        
        # Per-market sheets
        markets = set(lead.get('market', '') for lead in leads)
        for market in sorted(markets):
            if not market:
                continue
            market_leads = [l for l in leads if l.get('market') == market]
            if market_leads:
                ws_market = wb.create_sheet(market.upper())
                self._write_leads_to_sheet(ws_market, market_leads, headers, widths)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        self._write_summary(ws_summary, leads)
        
        wb.save(filepath)
        return filepath
    
    def _write_leads_to_sheet(self, ws, leads: List[Dict], headers: List[str], widths: List[int]):
        """Write leads to a worksheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        for row_idx, lead in enumerate(leads, 2):
            ws.cell(row=row_idx, column=1, value=lead.get('name', ''))
            ws.cell(row=row_idx, column=2, value=lead.get('title', ''))
            ws.cell(row=row_idx, column=3, value=lead.get('icp_category', ''))
            ws.cell(row=row_idx, column=4, value=lead.get('company', ''))
            ws.cell(row=row_idx, column=5, value=lead.get('company_type', ''))
            ws.cell(row=row_idx, column=6, value=lead.get('linkedin_url', ''))
            ws.cell(row=row_idx, column=7, value=lead.get('email', ''))
            ws.cell(row=row_idx, column=8, value=lead.get('market', '').upper())
            ws.cell(row=row_idx, column=9, value=lead.get('priority_score', 0))
            ws.cell(row=row_idx, column=10, value=lead.get('top_shows', ''))
            ws.cell(row=row_idx, column=11, value=lead.get('catalog_size', ''))
            ws.cell(row=row_idx, column=12, value=lead.get('location', ''))
        
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _write_summary(self, ws, leads: List[Dict]):
        """Write summary sheet."""
        ws['A1'] = "TrueSync Lead Generation Summary"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A4'] = f"Total Leads: {len(leads)}"
        ws['A5'] = f"Leads with Email: {sum(1 for l in leads if l.get('email'))}"
        
        # ICP breakdown
        ws['A7'] = "By ICP Category:"
        ws['A7'].font = Font(bold=True)
        
        icp_counts = {}
        for lead in leads:
            icp = lead.get('icp_category', 'Other')
            icp_counts[icp] = icp_counts.get(icp, 0) + 1
        
        row = 8
        for icp, count in sorted(icp_counts.items()):
            ws.cell(row=row, column=1, value=f"  {icp}: {count}")
            row += 1
        
        # Market breakdown
        row += 1
        ws.cell(row=row, column=1, value="By Market:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        market_counts = {}
        for lead in leads:
            market = lead.get('market', 'Unknown').upper()
            market_counts[market] = market_counts.get(market, 0) + 1
        
        for market, count in sorted(market_counts.items()):
            emails = sum(1 for l in leads if l.get('market', '').upper() == market and l.get('email'))
            ws.cell(row=row, column=1, value=f"  {market}: {count} leads, {emails} emails")
            row += 1
        
        # Company breakdown
        row += 1
        ws.cell(row=row, column=1, value="By Company:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        companies = set(lead.get('company', '') for lead in leads)
        for company in sorted(companies):
            if company:
                count = sum(1 for l in leads if l.get('company') == company)
                emails = sum(1 for l in leads if l.get('company') == company and l.get('email'))
                ws.cell(row=row, column=1, value=f"  {company}: {count} leads, {emails} emails")
                row += 1
        
        ws.column_dimensions['A'].width = 60


def main():
    parser = argparse.ArgumentParser(description='TrueSync Lead Generation Pipeline')
    parser.add_argument('--tier1', action='store_true', help='Run Tier 1 markets (UK, USA, Spain)')
    parser.add_argument('--tier2', action='store_true', help='Run Tier 2 markets (Germany, France, Korea)')
    parser.add_argument('--avod', action='store_true', help='Run AVOD platforms only')
    parser.add_argument('--all', action='store_true', help='Run all companies')
    parser.add_argument('--company', type=str, help='Run specific company by name')
    parser.add_argument('--leads', type=int, default=10, help='Max leads per company')
    
    args = parser.parse_args()
    
    # Load config
    config = load_config()
    roles = get_target_roles(config)
    
    print("\n" + "="*70)
    print("TRUESYNC LEAD GENERATION - ICP TARGETING")
    print("="*70)
    
    # Determine which companies to process
    if args.company:
        companies = [c for c in COMPANIES if args.company.lower() in c['name'].lower()]
        if not companies:
            print(f"Company '{args.company}' not found!")
            sys.exit(1)
    elif args.tier1:
        companies = get_tier1_companies()
    elif args.tier2:
        companies = get_tier2_companies()
    elif args.avod:
        companies = get_avod_companies()
    elif args.all:
        companies = get_all_companies()
    else:
        # Default: test with Netflix only
        companies = [c for c in COMPANIES if c['name'] == 'Netflix US']
        if not companies:
            companies = [COMPANIES[0]]  # First company as fallback
    
    print(f"\nTarget Companies: {len(companies)}")
    for c in companies[:5]:
        print(f"  - {c['name']} ({c['market'].upper()})")
    if len(companies) > 5:
        print(f"  ... and {len(companies) - 5} more")
    
    print(f"\nTarget ICP Roles ({len(roles)} total):")
    for role in roles[:5]:
        print(f"  - {role}")
    if len(roles) > 5:
        print(f"  ... and {len(roles) - 5} more")
    
    print(f"\nMax leads per company: {args.leads}")
    
    # Run pipeline
    generator = LinkedInLeadGenerator(config)
    leads = generator.run_batch(
        companies=companies,
        roles=roles,
        leads_per_company=args.leads
    )
    
    print(f"\n{'='*70}")
    print("RESULTS SUMMARY")
    print('='*70)
    print(f"Total leads found: {len(leads)}")
    print(f"Leads with email: {sum(1 for l in leads if l.get('email'))}")
    
    # ICP breakdown
    icp_counts = {}
    for lead in leads:
        icp = lead.get('icp_category', 'Other')
        icp_counts[icp] = icp_counts.get(icp, 0) + 1
    
    print("\nBy ICP Category:")
    for icp, count in sorted(icp_counts.items()):
        print(f"  {icp}: {count}")
    
    # Export to Excel
    if leads:
        exporter = ExcelExporter()
        filepath = exporter.export(leads)
        print(f"\n✓ Exported to: {filepath}")
        
        # Print sample
        print(f"\n{'='*70}")
        print("TOP LEADS:")
        print('='*70)
        for i, lead in enumerate(leads[:5], 1):
            print(f"\n{i}. {lead.get('name', 'N/A')}")
            print(f"   Title: {lead.get('title', 'N/A')}")
            print(f"   ICP: {lead.get('icp_category', 'N/A')}")
            print(f"   Company: {lead.get('company', 'N/A')}")
            print(f"   Email: {lead.get('email', 'NOT FOUND')}")
            print(f"   Score: {lead.get('priority_score', 0)}")
    else:
        print("\nNo leads found!")
    
    print("\n" + "="*70)
    print("PIPELINE COMPLETE")
    print("="*70 + "\n")


if __name__ == '__main__':
    main()
